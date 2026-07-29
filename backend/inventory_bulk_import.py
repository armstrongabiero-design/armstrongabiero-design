"""Excel template and parsing for bulk inventory item import (upsert by SKU)."""
from __future__ import annotations

import io
import re
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from country_utils import normalize_country_code
from models.enums import CurrencyEnum

BULK_HEADERS = [
    "SKU",
    "Name",
    "Category",
    "Country",
    "Location",
    "Quantity",
    "Reorder Level",
    "Unit Cost",
    "Currency",
    "Lead Time Days",
]

_HEADER_ALIASES = {
    "sku": "sku",
    "name": "name",
    "item name": "name",
    "category": "category",
    "country": "country",
    "location": "location",
    "quantity": "quantity",
    "qty": "quantity",
    "reorder level": "reorder_level",
    "reorder": "reorder_level",
    "unit cost": "unit_cost",
    "cost": "unit_cost",
    "currency": "currency",
    "lead time days": "lead_time_days",
    "lead time": "lead_time_days",
}

SAMPLE_ROW = [
    "FLT-OIL-5W30",
    "Engine Oil 5W-30",
    "Lubricants",
    "GH",
    "Accra Warehouse",
    24,
    10,
    85.0,
    "GHS",
    7,
]

_NUMBERED_INSTRUCTION = re.compile(r"^\d+\.\s")


def build_inventory_template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")

    for col, title in enumerate(BULK_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font

    for col, value in enumerate(SAMPLE_ROW, start=1):
        ws.cell(row=2, column=col, value=value)

    for col in range(1, len(BULK_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    guide = wb.create_sheet("Instructions")
    guide["A1"] = "Inventory bulk upload — how to use"
    guide["A1"].font = Font(bold=True)
    guide["A3"] = "1. Keep row 1 headers unchanged. Add one item per row from row 3."
    guide["A4"] = "2. SKU is the unique key — existing items with the same SKU are updated."
    guide["A5"] = "3. Required: SKU, Name, Category, Location, Unit Cost, Currency."
    guide["A6"] = "4. Country: ISO alpha-2. Blank rows use the country selected in the upload dialog."
    guide["A7"] = "5. Quantity defaults to 0; Reorder Level to 10; Lead Time Days to 7 if blank."
    guide["A8"] = "6. Currency: GHS, LRD, USD, or STN."
    guide["A9"] = "7. Delete the sample row before uploading."
    guide.column_dimensions["A"].width = 88

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _normalize_header(value: Any) -> Optional[str]:
    if value is None:
        return None
    return _HEADER_ALIASES.get(str(value).strip().lower())


def _parse_int(value: Any, field: str, *, required: bool = False, default: Optional[int] = None) -> Optional[int]:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        if required:
            raise ValueError(f"{field} is required")
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid {field}: {value}") from exc


def _parse_float(value: Any, field: str, *, required: bool = True) -> Optional[float]:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        if required:
            raise ValueError(f"{field} is required")
        return None
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid {field}: {value}") from exc


def _normalize_currency(value: Any) -> str:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        raise ValueError("Currency is required")
    text = str(value).strip().upper()
    valid = {c.value for c in CurrencyEnum}
    if text not in valid:
        raise ValueError(f"Currency must be one of: {', '.join(sorted(valid))}")
    return text


def parse_inventory_bulk_upload(
    file_bytes: bytes,
    default_country: Optional[str] = None,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if name.strip().lower() in ("inventory", "items", "sheet1"):
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ValueError("Spreadsheet is empty") from exc

    headers: List[Optional[str]] = []
    for cell in header_row:
        if cell is None:
            headers.append(None)
            continue
        text = str(cell).strip()
        if _NUMBERED_INSTRUCTION.match(text):
            headers.append(None)
            continue
        headers.append(_normalize_header(text))

    if "sku" not in headers or "name" not in headers:
        raise ValueError("Template must include SKU and Name columns")

    rows: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []

    for excel_row_idx, values in enumerate(rows_iter, start=2):
        if not values or all(v is None or (isinstance(v, str) and not str(v).strip()) for v in values):
            continue

        mapped: Dict[str, Any] = {}
        for idx, key in enumerate(headers):
            if key and idx < len(values):
                mapped[key] = values[idx]

        # Skip sample-looking empty after instruction sheet confusion
        sku = str(mapped.get("sku") or "").strip()
        if not sku:
            errors.append({"row": excel_row_idx, "message": "SKU is required"})
            continue

        try:
            country_raw = mapped.get("country")
            if country_raw is None or (isinstance(country_raw, str) and not str(country_raw).strip()):
                if not default_country:
                    raise ValueError("Country is required (set per row or in the upload dialog)")
                country = default_country
            else:
                country = normalize_country_code(str(country_raw).strip())

            name = str(mapped.get("name") or "").strip()
            if not name:
                raise ValueError("Name is required")
            category = str(mapped.get("category") or "").strip()
            if not category:
                raise ValueError("Category is required")
            location = str(mapped.get("location") or "").strip()
            if not location:
                raise ValueError("Location is required")

            row = {
                "row": excel_row_idx,
                "sku": sku,
                "name": name,
                "category": category,
                "country": country,
                "location": location,
                "quantity": _parse_int(mapped.get("quantity"), "Quantity", default=0) or 0,
                "reorder_level": _parse_int(mapped.get("reorder_level"), "Reorder Level", default=10) or 10,
                "unit_cost": _parse_float(mapped.get("unit_cost"), "Unit Cost"),
                "currency": _normalize_currency(mapped.get("currency")),
                "lead_time_days": _parse_int(mapped.get("lead_time_days"), "Lead Time Days", default=7) or 7,
            }
            rows.append(row)
        except Exception as exc:
            errors.append({"row": excel_row_idx, "sku": sku, "message": str(exc)})

    return rows, errors
