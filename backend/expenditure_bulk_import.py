"""Excel template and parsing for bulk expenditure (expense) import."""
from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from country_utils import normalize_country_code
from models.enums import CurrencyEnum

BULK_HEADERS = [
    "Date",
    "Country",
    "Category",
    "Description",
    "Amount",
    "Currency",
    "Vehicle Registration",
    "Driver License Number",
]

_HEADER_ALIASES = {
    "date": "date",
    "expense date": "date",
    "country": "country",
    "category": "category",
    "description": "description",
    "amount": "amount",
    "cost": "amount",
    "currency": "currency",
    "vehicle registration": "registration_number",
    "registration number": "registration_number",
    "vehicle": "registration_number",
    "driver license number": "license_number",
    "license number": "license_number",
    "driver license": "license_number",
}

SAMPLE_ROW = [
    "2026-07-15",
    "GH",
    "Tolls",
    "Accra–Tema highway tolls",
    25.0,
    "GHS",
    "GR-1234-20",
    "",
]

_NUMBERED_INSTRUCTION = re.compile(r"^\d+\.\s")


def build_expenditure_template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenditures"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")

    for col, title in enumerate(BULK_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font

    for col, value in enumerate(SAMPLE_ROW, start=1):
        ws.cell(row=2, column=col, value=value)

    for col in range(1, len(BULK_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    guide = wb.create_sheet("Instructions")
    guide["A1"] = "Expenditure / expense bulk upload — how to use"
    guide["A1"].font = Font(bold=True)
    guide["A3"] = "1. Keep row 1 headers unchanged. Add one expense per row from row 3."
    guide["A4"] = "2. Required: Date, Category, Description, Amount, Currency."
    guide["A5"] = "3. Country: ISO alpha-2. Blank rows use the country selected in the upload dialog."
    guide["A6"] = "4. Dates: YYYY-MM-DD. Currency: GHS, LRD, USD, or STN."
    guide["A7"] = "5. Vehicle Registration and Driver License Number are optional; must match existing records if set."
    guide["A8"] = "6. Each row creates a new expenditure (no upsert)."
    guide["A9"] = "7. Delete the sample row before uploading."
    guide.column_dimensions["A"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _normalize_header(value: Any) -> Optional[str]:
    if value is None:
        return None
    return _HEADER_ALIASES.get(str(value).strip().lower())


def _parse_date(value: Any) -> datetime:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        raise ValueError("Date is required")
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if hasattr(value, "year") and hasattr(value, "month") and not hasattr(value, "hour"):
        return datetime(value.year, value.month, value.day)
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError as exc:
        raise ValueError(f"Invalid Date: {value}") from exc


def _parse_float(value: Any, field: str) -> float:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        raise ValueError(f"{field} is required")
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid {field}: {value}") from exc


def _normalize_currency(value: Any) -> str:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        raise ValueError("Currency is required")
    text = str(value).strip().upper()
    valid = {c.value for c in CurrencyEnum}
    if text not in valid:
        raise ValueError(f"Currency must be one of: {', '.join(sorted(valid))}")
    return text


def parse_expenditure_bulk_upload(
    file_bytes: bytes,
    default_country: Optional[str] = None,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if name.strip().lower() in ("expenditures", "expenses", "expenditure", "sheet1"):
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ValueError("Spreadsheet is empty") from exc

    headers: List[Optional[str]] = []
    for cell in header_row:
        if cell is None:
            headers.append(None)
            continue
        text = str(cell).strip()
        if _NUMBERED_INSTRUCTION.match(text):
            headers.append(None)
            continue
        headers.append(_normalize_header(text))

    if "description" not in headers or "amount" not in headers:
        raise ValueError("Template must include Description and Amount columns")

    rows: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []

    for excel_row_idx, values in enumerate(rows_iter, start=2):
        if not values or all(v is None or (isinstance(v, str) and not str(v).strip()) for v in values):
            continue

        mapped: Dict[str, Any] = {}
        for idx, key in enumerate(headers):
            if key and idx < len(values):
                mapped[key] = values[idx]

        try:
            country_raw = mapped.get("country")
            if country_raw is None or (isinstance(country_raw, str) and not str(country_raw).strip()):
                if not default_country:
                    raise ValueError("Country is required (set per row or in the upload dialog)")
                country = default_country
            else:
                country = normalize_country_code(str(country_raw).strip())

            category = str(mapped.get("category") or "").strip()
            if not category:
                raise ValueError("Category is required")
            description = str(mapped.get("description") or "").strip()
            if not description:
                raise ValueError("Description is required")

            reg = str(mapped.get("registration_number") or "").strip() or None
            license_number = str(mapped.get("license_number") or "").strip() or None

            rows.append({
                "row": excel_row_idx,
                "date": _parse_date(mapped.get("date")),
                "country": country,
                "category": category,
                "description": description,
                "amount": _parse_float(mapped.get("amount"), "Amount"),
                "currency": _normalize_currency(mapped.get("currency")),
                "registration_number": reg,
                "license_number": license_number,
            })
        except Exception as exc:
            errors.append({"row": excel_row_idx, "message": str(exc)})

    return rows, errors
