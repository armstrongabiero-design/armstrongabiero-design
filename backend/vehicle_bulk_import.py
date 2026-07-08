"""Excel template and parsing for bulk vehicle import."""
from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from models.enums import CurrencyEnum
from models.fleet import VehicleCreate

# User-facing column headers (exact order in template)
BULK_HEADERS = [
    "Registration Number",
    "Make",
    "Model",
    "Year",
    "VIN",
    "Acquisition Date",
    "Odometer Reading (km)",
    "Acquisition Cost",
    "Currency",
]

_HEADER_ALIASES = {
    "registration number": "registration_number",
    "make": "make",
    "model": "model",
    "year": "year",
    "vin": "vin",
    "acquisition date": "acquisition_date",
    "odometer reading (km)": "odometer_reading",
    "odometer reading": "odometer_reading",
    "acquisition cost": "acquisition_cost",
    "currency": "acquisition_currency",
}

SAMPLE_ROW = [
    "GR-1234-20",
    "Toyota",
    "Hilux",
    2022,
    "JTFRB502200123456",
    "2022-06-15",
    45000,
    185000,
    "GHS",
]


def build_template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicles"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")

    for col, title in enumerate(BULK_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font

    for col, value in enumerate(SAMPLE_ROW, start=1):
        ws.cell(row=2, column=col, value=value)

    for col in range(1, len(BULK_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    guide = wb.create_sheet("Instructions")
    guide["A1"] = "Vehicle bulk upload — how to use"
    guide["A1"].font = Font(bold=True)
    guide["A3"] = "1. On the Vehicles sheet, keep row 1 (headers) unchanged."
    guide["A4"] = "2. Add one vehicle per row starting at row 3."
    guide["A5"] = "3. Acquisition Date format: YYYY-MM-DD (e.g. 2022-06-15)."
    guide["A6"] = "4. Currency: GHS, LRD, USD, or STN."
    guide["A7"] = "5. Delete the sample row (row 2) before uploading your data."
    guide["A8"] = "6. Select the fleet country in the upload screen (applies to all rows)."
    guide.column_dimensions["A"].width = 72

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _normalize_header(value: Any) -> Optional[str]:
    if value is None:
        return None
    key = str(value).strip().lower()
    return _HEADER_ALIASES.get(key)


def _parse_date(value: Any) -> datetime:
    if value is None or (isinstance(value, str) and not value.strip()):
        raise ValueError("Acquisition Date is required")
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if hasattr(value, "year") and hasattr(value, "month"):
        # openpyxl may return date objects
        return datetime(value.year, value.month, value.day)
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError as exc:
        raise ValueError(f"Invalid Acquisition Date: {value}") from exc


def _parse_number(value: Any, field: str, required: bool = True) -> float:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        if required:
            raise ValueError(f"{field} is required")
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid {field}: {value}") from exc


def _parse_int(value: Any, field: str) -> int:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        raise ValueError(f"{field} is required")
    try:
        return int(float(value))
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid {field}: {value}") from exc


def _parse_currency(value: Any) -> CurrencyEnum:
    if value is None or not str(value).strip():
        raise ValueError("Currency is required")
    code = str(value).strip().upper()
    try:
        return CurrencyEnum(code)
    except ValueError as exc:
        raise ValueError(f"Currency must be one of GHS, LRD, USD, STN (got {value})") from exc


def _row_is_empty(values: List[Any]) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in values)


def _cell_value(row: tuple, column_map: Dict[int, str], field: str) -> Any:
    for col_idx, name in column_map.items():
        if name == field:
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None and isinstance(val, str):
                return val.strip()
            return val
    return None


_NUMBERED_INSTRUCTION = re.compile(r"^\d+\.\s")
_SKIP_REGISTRATION_PREFIXES = ("instructions", "instruction:")


def _should_skip_row(row: tuple, column_map: Dict[int, str]) -> bool:
    """Ignore blank rows, instruction text, and rows with only column A filled."""
    if _row_is_empty(list(row)):
        return True

    reg = _cell_value(row, column_map, "registration_number")
    if reg is None or (isinstance(reg, str) and not str(reg).strip()):
        return True

    reg_text = str(reg).strip()
    reg_lower = reg_text.lower()
    if reg_lower.startswith(_SKIP_REGISTRATION_PREFIXES):
        return True
    if _NUMBERED_INSTRUCTION.match(reg_text):
        return True

    # Real data rows have values beyond registration (make, model, or year at minimum)
    make = _cell_value(row, column_map, "make")
    model = _cell_value(row, column_map, "model")
    year = _cell_value(row, column_map, "year")
    has_make = make is not None and str(make).strip()
    has_model = model is not None and str(model).strip()
    has_year = year is not None and str(year).strip()
    if not (has_make or has_model or has_year):
        return True

    return False


def parse_bulk_upload(
    file_bytes: bytes,
    *,
    country: str,
) -> Tuple[List[VehicleCreate], List[Dict[str, Any]]]:
    """Return (valid rows, errors with row numbers)."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb["Vehicles"] if "Vehicles" in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The spreadsheet is empty")

    header_row = rows[0]
    column_map: Dict[int, str] = {}
    for idx, header in enumerate(header_row):
        field = _normalize_header(header)
        if field:
            column_map[idx] = field

    present = set(column_map.values())
    missing_labels = [
        h for h in BULK_HEADERS if _HEADER_ALIASES.get(h.lower()) not in present
    ]
    if missing_labels:
        raise ValueError(f"Missing required columns: {', '.join(missing_labels)}")

    creates: List[VehicleCreate] = []
    errors: List[Dict[str, Any]] = []
    seen_registrations: set[str] = set()

    for row_idx, row in enumerate(rows[1:], start=2):
        if _should_skip_row(row, column_map):
            continue

        data: Dict[str, Any] = {"country": country}
        try:
            for col_idx, field in column_map.items():
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None and isinstance(val, str):
                    val = val.strip()
                data[field] = val

            reg = str(data.get("registration_number") or "").strip()
            if not reg:
                raise ValueError("Registration Number is required")

            reg_key = reg.upper()
            if reg_key in seen_registrations:
                raise ValueError(f"Duplicate registration in file: {reg}")
            seen_registrations.add(reg_key)

            vehicle_input = VehicleCreate(
                country=country,
                registration_number=reg,
                make=str(data.get("make") or "").strip(),
                model=str(data.get("model") or "").strip(),
                year=_parse_int(data.get("year"), "Year"),
                vin=str(data.get("vin") or "").strip(),
                acquisition_date=_parse_date(data.get("acquisition_date")),
                odometer_reading=_parse_number(data.get("odometer_reading"), "Odometer Reading (km)", required=False),
                acquisition_cost=_parse_number(data.get("acquisition_cost"), "Acquisition Cost"),
                acquisition_currency=_parse_currency(data.get("acquisition_currency")),
            )

            if not vehicle_input.make:
                raise ValueError("Make is required")
            if not vehicle_input.model:
                raise ValueError("Model is required")
            if not vehicle_input.vin:
                raise ValueError("VIN is required")

            creates.append(vehicle_input)
        except ValueError as exc:
            errors.append({
                "row": row_idx,
                "registration_number": str(data.get("registration_number") or ""),
                "message": str(exc),
            })

    wb.close()
    return creates, errors
