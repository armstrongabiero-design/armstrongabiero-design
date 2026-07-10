"""Excel template and parsing for bulk logbook import."""
from __future__ import annotations

import io
import re
from datetime import datetime, time
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

BULK_HEADERS = [
    "License Number",
    "Vehicle Registration",
    "Date",
    "Start Time",
    "End Time",
    "Start Location",
    "End Location",
    "Start Odometer",
    "End Odometer",
    "Purpose",
    "Fuel Used (L)",
    "Notes",
]

_HEADER_ALIASES = {
    "license number": "license_number",
    "vehicle registration": "registration_number",
    "registration number": "registration_number",
    "date": "date",
    "start time": "start_time",
    "end time": "end_time",
    "start location": "start_location",
    "end location": "end_location",
    "start odometer": "start_odometer",
    "end odometer": "end_odometer",
    "purpose": "purpose",
    "fuel used (l)": "fuel_used_liters",
    "fuel used": "fuel_used_liters",
    "notes": "notes",
}

SAMPLE_ROW = [
    "GH-DL-123456",
    "GR-1234-20",
    "2026-07-01",
    "08:00",
    "12:30",
    "Accra Depot",
    "Tema Port",
    45000,
    45120,
    "Delivery run",
    18.5,
    "",
]

_NUMBERED_INSTRUCTION = re.compile(r"^\d+\.\s")


def build_logbook_template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Logbook"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")

    for col, title in enumerate(BULK_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font

    for col, value in enumerate(SAMPLE_ROW, start=1):
        ws.cell(row=2, column=col, value=value)

    for col in range(1, len(BULK_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    guide = wb.create_sheet("Instructions")
    guide["A1"] = "Logbook bulk upload — how to use"
    guide["A1"].font = Font(bold=True)
    guide["A3"] = "1. Keep row 1 headers unchanged. Add one trip per row from row 3."
    guide["A4"] = "2. Match drivers by License Number and vehicles by Registration Number."
    guide["A5"] = "3. Date: YYYY-MM-DD. Times: HH:MM (24-hour)."
    guide["A6"] = "4. A completed pre-trip checklist is required for that driver, vehicle, and date."
    guide["A7"] = "5. Delete the sample row before uploading."
    guide.column_dimensions["A"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _normalize_header(value: Any) -> Optional[str]:
    if value is None:
        return None
    return _HEADER_ALIASES.get(str(value).strip().lower())


def _parse_date(value: Any) -> datetime:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        raise ValueError("Date is required")
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if hasattr(value, "year") and hasattr(value, "month") and not hasattr(value, "hour"):
        return datetime(value.year, value.month, value.day)
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError as exc:
        raise ValueError(f"Invalid Date: {value}") from exc


def _parse_time_on_date(value: Any, base_date: datetime, field: str, required: bool = True) -> Optional[datetime]:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        if required:
            raise ValueError(f"{field} is required")
        return None
    if isinstance(value, datetime):
        return datetime.combine(base_date.date(), value.time())
    if isinstance(value, time):
        return datetime.combine(base_date.date(), value)
    text = str(value).strip()
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M%p"):
        try:
            parsed = datetime.strptime(text, fmt).time()
            return datetime.combine(base_date.date(), parsed)
        except ValueError:
            continue
    raise ValueError(f"Invalid {field}: {value}")


def _parse_float(value: Any, field: str, required: bool = True) -> Optional[float]:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        if required:
            raise ValueError(f"{field} is required")
        return None
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid {field}: {value}") from exc


def _row_is_empty(values: List[Any]) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in values)


def _cell_value(row: tuple, column_map: Dict[int, str], field: str) -> Any:
    for col_idx, name in column_map.items():
        if name == field:
            val = row[col_idx] if col_idx < len(row) else None
            if isinstance(val, str):
                return val.strip()
            return val
    return None


def _should_skip_row(row: tuple, column_map: Dict[int, str]) -> bool:
    if _row_is_empty(list(row)):
        return True
    license_no = _cell_value(row, column_map, "license_number")
    if license_no is None or (isinstance(license_no, str) and not str(license_no).strip()):
        return True
    text = str(license_no).strip()
    if text.lower().startswith("instructions") or _NUMBERED_INSTRUCTION.match(text):
        return True
    return False


def parse_logbook_bulk_upload(file_bytes: bytes) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Return (parsed rows with license/registration keys, parse errors)."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb["Logbook"] if "Logbook" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The spreadsheet is empty")

    column_map: Dict[int, str] = {}
    for idx, header in enumerate(rows[0]):
        field = _normalize_header(header)
        if field:
            column_map[idx] = field

    present = set(column_map.values())
    required = {
        "license_number", "registration_number", "date", "start_time",
        "start_location", "start_odometer", "purpose",
    }
    missing = [
        h for h in BULK_HEADERS
        if _HEADER_ALIASES.get(h.lower()) in required and _HEADER_ALIASES.get(h.lower()) not in present
    ]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    parsed: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        if _should_skip_row(row, column_map):
            continue

        data: Dict[str, Any] = {"row": row_idx}
        try:
            for col_idx, field in column_map.items():
                val = row[col_idx] if col_idx < len(row) else None
                if isinstance(val, str):
                    val = val.strip()
                data[field] = val

            license_no = str(data.get("license_number") or "").strip()
            registration = str(data.get("registration_number") or "").strip()
            if not license_no:
                raise ValueError("License Number is required")
            if not registration:
                raise ValueError("Vehicle Registration is required")

            trip_date = _parse_date(data.get("date"))
            start_time = _parse_time_on_date(data.get("start_time"), trip_date, "Start Time", required=True)
            end_time = _parse_time_on_date(data.get("end_time"), trip_date, "End Time", required=False)
            start_location = str(data.get("start_location") or "").strip()
            end_location = str(data.get("end_location") or "").strip() or None
            purpose = str(data.get("purpose") or "").strip()
            notes = str(data.get("notes") or "").strip() or None
            start_odo = _parse_float(data.get("start_odometer"), "Start Odometer")
            end_odo = _parse_float(data.get("end_odometer"), "End Odometer", required=False)
            fuel = _parse_float(data.get("fuel_used_liters"), "Fuel Used (L)", required=False)

            if not start_location:
                raise ValueError("Start Location is required")
            if not purpose:
                raise ValueError("Purpose is required")

            parsed.append({
                "row": row_idx,
                "license_number": license_no,
                "registration_number": registration,
                "date": trip_date,
                "start_time": start_time,
                "end_time": end_time,
                "start_location": start_location,
                "end_location": end_location,
                "start_odometer": start_odo,
                "end_odometer": end_odo,
                "purpose": purpose,
                "fuel_used_liters": fuel,
                "notes": notes,
            })
        except ValueError as exc:
            errors.append({
                "row": row_idx,
                "license_number": str(data.get("license_number") or ""),
                "registration_number": str(data.get("registration_number") or ""),
                "message": str(exc),
            })

    wb.close()
    return parsed, errors
