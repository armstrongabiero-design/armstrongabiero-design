"""Excel template and parsing for bulk maintenance record import."""
from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from models.enums import CurrencyEnum, MaintenanceType, WorkStatus

BULK_HEADERS = [
    "Vehicle Registration",
    "Maintenance Type",
    "Description",
    "Scheduled Date",
    "Next Service Date",
    "Odometer",
    "Next Service Odometer",
    "Cost",
    "Currency",
    "Work Status",
    "Notes",
]

_HEADER_ALIASES = {
    "vehicle registration": "registration_number",
    "vehicle": "registration_number",
    "registration number": "registration_number",
    "maintenance type": "maintenance_type",
    "type": "maintenance_type",
    "description": "description",
    "scheduled date": "scheduled_date",
    "next service date": "next_due_date",
    "next due date": "next_due_date",
    "odometer": "odometer_at_maintenance",
    "odometer at maintenance": "odometer_at_maintenance",
    "odometer at service": "odometer_at_maintenance",
    "next service odometer": "next_service_odometer",
    "next service odo": "next_service_odometer",
    "cost": "cost",
    "currency": "currency",
    "work status": "work_status",
    "notes": "notes",
}

_TYPE_ALIASES = {
    "predictive": MaintenanceType.PREDICTIVE.value,
    "corrective": MaintenanceType.CORRECTIVE.value,
    "routine": MaintenanceType.ROUTINE.value,
    "scheduled": MaintenanceType.ROUTINE.value,
    "unscheduled": MaintenanceType.CORRECTIVE.value,
}

_WORK_STATUS_ALIASES = {
    "work in progress": WorkStatus.WORK_IN_PROGRESS.value,
    "work_in_progress": WorkStatus.WORK_IN_PROGRESS.value,
    "in progress": WorkStatus.WORK_IN_PROGRESS.value,
    "work completed": WorkStatus.WORK_COMPLETED.value,
    "work_completed": WorkStatus.WORK_COMPLETED.value,
    "completed": WorkStatus.WORK_COMPLETED.value,
    "etc": WorkStatus.ETC.value,
    "estimated time of completion": WorkStatus.ETC.value,
    "additional work required": WorkStatus.ADDITIONAL_WORK_REQUIRED.value,
    "additional_work_required": WorkStatus.ADDITIONAL_WORK_REQUIRED.value,
}

SAMPLE_ROW = [
    "GR-1234-20",
    "ROUTINE",
    "Oil change and filter replacement",
    "2026-07-01",
    "",
    45000,
    "",
    350.0,
    "GHS",
    "WORK_IN_PROGRESS",
    "",
]

_NUMBERED_INSTRUCTION = re.compile(r"^\d+\.\s")


def build_maintenance_template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Maintenance"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")

    for col, title in enumerate(BULK_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font

    for col, value in enumerate(SAMPLE_ROW, start=1):
        ws.cell(row=2, column=col, value=value)

    widths = [20, 16, 32, 14, 16, 12, 18, 10, 10, 18, 24]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    guide = wb.create_sheet("Instructions")
    guide["A1"] = "Maintenance bulk upload — how to use"
    guide["A1"].font = Font(bold=True)
    guide["A3"] = "1. Keep row 1 headers unchanged. Add one record per row from row 3."
    guide["A4"] = "2. Vehicle Registration must match an existing vehicle."
    guide["A5"] = "3. Maintenance Type: PREDICTIVE, CORRECTIVE, or ROUTINE."
    guide["A6"] = "4. Dates: YYYY-MM-DD. Currency: GHS, LRD, USD, or STN."
    guide["A7"] = "5. Next Service Date / Next Service Odometer are optional — auto-calculated from Master Data defaults when blank."
    guide["A8"] = "6. Work Status (optional): WORK_IN_PROGRESS, WORK_COMPLETED, ETC, ADDITIONAL_WORK_REQUIRED."
    guide["A9"] = "7. Delete the sample row before uploading."
    guide.column_dimensions["A"].width = 88

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _normalize_header(value: Any) -> Optional[str]:
    if value is None:
        return None
    return _HEADER_ALIASES.get(str(value).strip().lower())


def _parse_date(value: Any, field: str, required: bool = True) -> Optional[datetime]:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        if required:
            raise ValueError(f"{field} is required")
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if hasattr(value, "year") and hasattr(value, "month") and not hasattr(value, "hour"):
        return datetime(value.year, value.month, value.day)
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError as exc:
        raise ValueError(f"Invalid {field}: {value}") from exc


def _parse_float(value: Any, field: str, required: bool = True) -> Optional[float]:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        if required:
            raise ValueError(f"{field} is required")
        return None
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid {field}: {value}") from exc


def _normalize_type(value: Any) -> str:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        raise ValueError("Maintenance Type is required")
    text = str(value).strip()
    upper = text.upper().replace(" ", "_")
    valid = {t.value for t in MaintenanceType}
    if upper in valid:
        return upper
    aliased = _TYPE_ALIASES.get(text.lower())
    if aliased:
        return aliased
    raise ValueError(f"Unknown Maintenance Type: {value}")


def _normalize_work_status(value: Any) -> str:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        return WorkStatus.WORK_IN_PROGRESS.value
    text = str(value).strip()
    upper = text.upper().replace(" ", "_")
    valid = {s.value for s in WorkStatus}
    if upper in valid:
        return upper
    aliased = _WORK_STATUS_ALIASES.get(text.lower())
    if aliased:
        return aliased
    raise ValueError(f"Unknown Work Status: {value}")


def _normalize_currency(value: Any) -> str:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        raise ValueError("Currency is required")
    text = str(value).strip().upper()
    valid = {c.value for c in CurrencyEnum}
    if text not in valid:
        raise ValueError(f"Currency must be one of: {', '.join(sorted(valid))}")
    return text


def _row_is_empty(values: List[Any]) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in values)


def _cell_value(row: tuple, column_map: Dict[int, str], field: str) -> Any:
    for col_idx, name in column_map.items():
        if name == field:
            val = row[col_idx] if col_idx < len(row) else None
            if isinstance(val, str):
                return val.strip()
            return val
    return None


def _should_skip_row(row: tuple, column_map: Dict[int, str]) -> bool:
    if _row_is_empty(list(row)):
        return True
    reg = _cell_value(row, column_map, "registration_number")
    if reg is None or (isinstance(reg, str) and not str(reg).strip()):
        return True
    text = str(reg).strip()
    if text.lower().startswith("instructions") or _NUMBERED_INSTRUCTION.match(text):
        return True
    return False


def parse_maintenance_bulk_upload(file_bytes: bytes) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb["Maintenance"] if "Maintenance" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The spreadsheet is empty")

    column_map: Dict[int, str] = {}
    for idx, header in enumerate(rows[0]):
        field = _normalize_header(header)
        if field:
            column_map[idx] = field

    present = set(column_map.values())
    required = {
        "registration_number",
        "maintenance_type",
        "description",
        "scheduled_date",
        "odometer_at_maintenance",
        "cost",
        "currency",
    }
    missing = [field for field in required if field not in present]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

    parsed: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        if _should_skip_row(row, column_map):
            continue

        data: Dict[str, Any] = {"row": row_idx}
        try:
            for col_idx, field in column_map.items():
                val = row[col_idx] if col_idx < len(row) else None
                if isinstance(val, str):
                    val = val.strip()
                data[field] = val

            registration = str(data.get("registration_number") or "").strip()
            if not registration:
                raise ValueError("Vehicle Registration is required")

            description = str(data.get("description") or "").strip()
            if not description:
                raise ValueError("Description is required")

            notes = str(data.get("notes") or "").strip() or None
            next_odo = _parse_float(data.get("next_service_odometer"), "Next Service Odometer", required=False)

            parsed.append({
                "row": row_idx,
                "registration_number": registration,
                "maintenance_type": _normalize_type(data.get("maintenance_type")),
                "description": description,
                "scheduled_date": _parse_date(data.get("scheduled_date"), "Scheduled Date"),
                "odometer_at_maintenance": _parse_float(data.get("odometer_at_maintenance"), "Odometer"),
                "cost": _parse_float(data.get("cost"), "Cost"),
                "currency": _normalize_currency(data.get("currency")),
                "next_due_date": _parse_date(data.get("next_due_date"), "Next Service Date", required=False),
                "next_service_odometer": next_odo,
                "work_status": _normalize_work_status(data.get("work_status")),
                "notes": notes,
            })
        except ValueError as exc:
            errors.append({
                "row": row_idx,
                "registration_number": str(data.get("registration_number") or ""),
                "message": str(exc),
            })

    wb.close()
    return parsed, errors
