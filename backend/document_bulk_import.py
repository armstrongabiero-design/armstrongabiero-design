"""Excel template and parsing for bulk document import (metadata + file matching)."""
from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from models.enums import DocumentType

BULK_HEADERS = [
    "Filename",
    "Document Type",
    "Entity Type",
    "Registration or License Number",
    "Document Number",
    "Issue Date",
    "Expiry Date",
]

_HEADER_ALIASES = {
    "filename": "filename",
    "file name": "filename",
    "document type": "document_type",
    "entity type": "entity_type",
    "registration or license number": "entity_key",
    "registration number": "entity_key",
    "license number": "entity_key",
    "document number": "document_number",
    "issue date": "issue_date",
    "expiry date": "expiry_date",
}

_TYPE_ALIASES = {
    "roadworthy": DocumentType.ROADWORTHY_CERT.value,
    "roadworthy cert": DocumentType.ROADWORTHY_CERT.value,
    "roadworthy certificate": DocumentType.ROADWORTHY_CERT.value,
    "roadworthy_cert": DocumentType.ROADWORTHY_CERT.value,
    "ama sticker": DocumentType.AMA_STICKER.value,
    "ama stickers": DocumentType.AMA_STICKER.value,
    "ama_sticker": DocumentType.AMA_STICKER.value,
    "insurance": DocumentType.INSURANCE.value,
    "insurance documents": DocumentType.INSURANCE.value,
    "driver license": DocumentType.DRIVER_LICENSE.value,
    "drivers license": DocumentType.DRIVER_LICENSE.value,
    "driver's license": DocumentType.DRIVER_LICENSE.value,
    "drivers' licenses": DocumentType.DRIVER_LICENSE.value,
    "driver_license": DocumentType.DRIVER_LICENSE.value,
    "vehicle registration": DocumentType.VEHICLE_REGISTRATION.value,
    "vehicle registration certificate": DocumentType.VEHICLE_REGISTRATION.value,
    "vrc": DocumentType.VEHICLE_REGISTRATION.value,
    "vehicle_registration": DocumentType.VEHICLE_REGISTRATION.value,
    "other": DocumentType.OTHER.value,
    "other documents": DocumentType.OTHER.value,
}

_VEHICLE_TYPES = frozenset({
    DocumentType.ROADWORTHY_CERT.value,
    DocumentType.AMA_STICKER.value,
    DocumentType.INSURANCE.value,
    DocumentType.VEHICLE_REGISTRATION.value,
})

SAMPLE_ROW = [
    "roadworthy-gr1234.pdf",
    "ROADWORTHY_CERT",
    "VEHICLE",
    "GR-1234-20",
    "RW-2026-001",
    "2026-01-15",
    "2027-01-14",
]

_NUMBERED_INSTRUCTION = re.compile(r"^\d+\.\s")


def build_document_template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Documents"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")

    for col, title in enumerate(BULK_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font

    for col, value in enumerate(SAMPLE_ROW, start=1):
        ws.cell(row=2, column=col, value=value)

    for col in range(1, len(BULK_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 28

    guide = wb.create_sheet("Instructions")
    guide["A1"] = "Document bulk upload — how to use"
    guide["A1"].font = Font(bold=True)
    guide["A3"] = "1. Keep row 1 headers unchanged. Add one document per row from row 3."
    guide["A4"] = "2. Filename must match an uploaded file name exactly (including extension)."
    guide["A5"] = "3. Document Type: ROADWORTHY_CERT, AMA_STICKER, INSURANCE, DRIVER_LICENSE, VEHICLE_REGISTRATION, OTHER."
    guide["A6"] = "4. Entity Type: VEHICLE or DRIVER. DRIVER_LICENSE forces DRIVER; most vehicle docs force VEHICLE."
    guide["A7"] = "5. Registration or License Number must match an existing vehicle or driver."
    guide["A8"] = "6. Dates: YYYY-MM-DD. Country is chosen in the upload dialog."
    guide["A9"] = "7. Delete the sample row before uploading."
    guide.column_dimensions["A"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _normalize_header(value: Any) -> Optional[str]:
    if value is None:
        return None
    return _HEADER_ALIASES.get(str(value).strip().lower())


def _parse_date(value: Any, field: str) -> datetime:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        raise ValueError(f"{field} is required")
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if hasattr(value, "year") and hasattr(value, "month") and not hasattr(value, "hour"):
        return datetime(value.year, value.month, value.day)
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError as exc:
        raise ValueError(f"Invalid {field}: {value}") from exc


def _normalize_document_type(value: Any) -> str:
    if value is None or (isinstance(value, str) and not value.strip()):
        raise ValueError("Document Type is required")
    text = str(value).strip()
    upper = text.upper().replace(" ", "_").replace("-", "_")
    valid = {t.value for t in DocumentType}
    if upper in valid:
        return upper
    aliased = _TYPE_ALIASES.get(text.lower())
    if aliased:
        return aliased
    raise ValueError(f"Unknown Document Type: {value}")


def _resolve_entity_type(doc_type: str, entity_type_raw: Any) -> str:
    if doc_type == DocumentType.DRIVER_LICENSE.value:
        return "DRIVER"
    if doc_type in _VEHICLE_TYPES:
        return "VEHICLE"
    text = str(entity_type_raw or "").strip().upper()
    if text in ("VEHICLE", "DRIVER"):
        return text
    raise ValueError("Entity Type must be VEHICLE or DRIVER for OTHER documents")


def _row_is_empty(values: List[Any]) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in values)


def _cell_value(row: tuple, column_map: Dict[int, str], field: str) -> Any:
    for col_idx, name in column_map.items():
        if name == field:
            val = row[col_idx] if col_idx < len(row) else None
            if isinstance(val, str):
                return val.strip()
            return val
    return None


def _should_skip_row(row: tuple, column_map: Dict[int, str]) -> bool:
    if _row_is_empty(list(row)):
        return True
    filename = _cell_value(row, column_map, "filename")
    if filename is None or (isinstance(filename, str) and not str(filename).strip()):
        return True
    text = str(filename).strip()
    if text.lower().startswith("instructions") or _NUMBERED_INSTRUCTION.match(text):
        return True
    return False


def parse_document_bulk_upload(file_bytes: bytes) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Return (parsed rows, parse errors)."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb["Documents"] if "Documents" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The spreadsheet is empty")

    column_map: Dict[int, str] = {}
    for idx, header in enumerate(rows[0]):
        field = _normalize_header(header)
        if field:
            column_map[idx] = field

    present = set(column_map.values())
    required = {
        "filename", "document_type", "entity_key",
        "document_number", "issue_date", "expiry_date",
    }
    missing = [
        h for h in BULK_HEADERS
        if _HEADER_ALIASES.get(h.lower()) in required and _HEADER_ALIASES.get(h.lower()) not in present
    ]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    parsed: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        if _should_skip_row(row, column_map):
            continue

        data: Dict[str, Any] = {"row": row_idx}
        try:
            for col_idx, field in column_map.items():
                val = row[col_idx] if col_idx < len(row) else None
                if isinstance(val, str):
                    val = val.strip()
                data[field] = val

            filename = str(data.get("filename") or "").strip()
            if not filename:
                raise ValueError("Filename is required")

            doc_type = _normalize_document_type(data.get("document_type"))
            entity_type = _resolve_entity_type(doc_type, data.get("entity_type"))
            entity_key = str(data.get("entity_key") or "").strip()
            if not entity_key:
                raise ValueError("Registration or License Number is required")

            document_number = str(data.get("document_number") or "").strip()
            if not document_number:
                raise ValueError("Document Number is required")

            issue_date = _parse_date(data.get("issue_date"), "Issue Date")
            expiry_date = _parse_date(data.get("expiry_date"), "Expiry Date")

            parsed.append({
                "row": row_idx,
                "filename": filename,
                "document_type": doc_type,
                "entity_type": entity_type,
                "entity_key": entity_key,
                "document_number": document_number,
                "issue_date": issue_date,
                "expiry_date": expiry_date,
            })
        except ValueError as exc:
            errors.append({
                "row": row_idx,
                "filename": str(data.get("filename") or ""),
                "message": str(exc),
            })

    wb.close()
    return parsed, errors
