"""Excel template and parsing for bulk driver import."""
from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from models.fleet import DriverCreate

BULK_HEADERS = [
    "First Name",
    "Last Name",
    "License Number",
    "License Expiry",
    "Phone",
    "Email",
]

_HEADER_ALIASES = {
    "first name": "first_name",
    "last name": "last_name",
    "license number": "license_number",
    "license expiry": "license_expiry",
    "phone": "phone",
    "email": "email",
}

SAMPLE_ROW = [
    "Kwame",
    "Mensah",
    "GH-DL-123456",
    "2027-12-31",
    "+233201234567",
    "kwame.mensah@example.com",
]

_NUMBERED_INSTRUCTION = re.compile(r"^\d+\.\s")


def build_driver_template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Drivers"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")

    for col, title in enumerate(BULK_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font

    for col, value in enumerate(SAMPLE_ROW, start=1):
        ws.cell(row=2, column=col, value=value)

    for col in range(1, len(BULK_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    guide = wb.create_sheet("Instructions")
    guide["A1"] = "Driver bulk upload — how to use"
    guide["A1"].font = Font(bold=True)
    guide["A3"] = "1. On the Drivers sheet, keep row 1 (headers) unchanged."
    guide["A4"] = "2. Add one driver per row starting at row 3."
    guide["A5"] = "3. License Expiry format: YYYY-MM-DD (e.g. 2027-12-31)."
    guide["A6"] = "4. Email is optional."
    guide["A7"] = "5. Delete the sample row (row 2) before uploading your data."
    guide["A8"] = "6. Select the fleet country in the upload screen (applies to all rows)."
    guide.column_dimensions["A"].width = 72

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _normalize_header(value: Any) -> Optional[str]:
    if value is None:
        return None
    return _HEADER_ALIASES.get(str(value).strip().lower())


def _parse_date(value: Any, field: str = "License Expiry") -> datetime:
    if value is None or (isinstance(value, str) and not value.strip()):
        raise ValueError(f"{field} is required")
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if hasattr(value, "year") and hasattr(value, "month"):
        return datetime(value.year, value.month, value.day)
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError as exc:
        raise ValueError(f"Invalid {field}: {value}") from exc


def _row_is_empty(values: List[Any]) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in values)


def _cell_value(row: tuple, column_map: Dict[int, str], field: str) -> Any:
    for col_idx, name in column_map.items():
        if name == field:
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None and isinstance(val, str):
                return val.strip()
            return val
    return None


def _should_skip_row(row: tuple, column_map: Dict[int, str]) -> bool:
    if _row_is_empty(list(row)):
        return True
    first = _cell_value(row, column_map, "first_name")
    if first is None or (isinstance(first, str) and not str(first).strip()):
        return True
    text = str(first).strip()
    if text.lower().startswith("instructions") or _NUMBERED_INSTRUCTION.match(text):
        return True
    last = _cell_value(row, column_map, "last_name")
    license_no = _cell_value(row, column_map, "license_number")
    if not (last or license_no):
        return True
    return False


def parse_driver_bulk_upload(
    file_bytes: bytes,
    *,
    country: str,
) -> Tuple[List[DriverCreate], List[Dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb["Drivers"] if "Drivers" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The spreadsheet is empty")

    column_map: Dict[int, str] = {}
    for idx, header in enumerate(rows[0]):
        field = _normalize_header(header)
        if field:
            column_map[idx] = field

    present = set(column_map.values())
    required = {"first_name", "last_name", "license_number", "license_expiry", "phone"}
    missing_labels = [
        h for h in BULK_HEADERS
        if _HEADER_ALIASES.get(h.lower()) in required and _HEADER_ALIASES.get(h.lower()) not in present
    ]
    if missing_labels:
        raise ValueError(f"Missing required columns: {', '.join(missing_labels)}")

    creates: List[DriverCreate] = []
    errors: List[Dict[str, Any]] = []
    seen_licenses: set[str] = set()

    for row_idx, row in enumerate(rows[1:], start=2):
        if _should_skip_row(row, column_map):
            continue

        data: Dict[str, Any] = {"country": country}
        try:
            for col_idx, field in column_map.items():
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None and isinstance(val, str):
                    val = val.strip()
                data[field] = val

            first = str(data.get("first_name") or "").strip()
            last = str(data.get("last_name") or "").strip()
            license_no = str(data.get("license_number") or "").strip()
            phone = str(data.get("phone") or "").strip()
            email_raw = data.get("email")
            email = str(email_raw).strip() if email_raw else None

            if not first:
                raise ValueError("First Name is required")
            if not last:
                raise ValueError("Last Name is required")
            if not license_no:
                raise ValueError("License Number is required")
            if not phone:
                raise ValueError("Phone is required")

            license_key = license_no.upper()
            if license_key in seen_licenses:
                raise ValueError(f"Duplicate license number in file: {license_no}")
            seen_licenses.add(license_key)

            creates.append(
                DriverCreate(
                    country=country,
                    first_name=first,
                    last_name=last,
                    license_number=license_no,
                    license_expiry=_parse_date(data.get("license_expiry")),
                    phone=phone,
                    email=email or None,
                )
            )
        except ValueError as exc:
            errors.append({
                "row": row_idx,
                "license_number": str(data.get("license_number") or ""),
                "message": str(exc),
            })

    wb.close()
    return creates, errors
