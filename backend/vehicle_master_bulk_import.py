"""Bulk import vehicle master records from Excel template."""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from models.vehicle_master import VEHICLE_MASTER_COLUMNS

HEADER_ROW = 1
DATA_START = 2

_NUMERIC_KEYS = {
    "year_of_manufacture", "quantity", "seating_capacity", "max_speed",
    "number_of_wheels", "engine_capacity_cc", "power_value", "cylinders", "book_value",
}


def build_vehicle_master_template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicle Master"
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(bold=True, color="FFFFFF")
    headers = [label for _, label in VEHICLE_MASTER_COLUMNS]
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = fill
        cell.font = font
        ws.column_dimensions[get_column_letter(col)].width = max(14, min(28, len(name) + 2))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _norm_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("/", " ").replace("  ", " ")


def _header_map(headers: List[Any]) -> Dict[str, int]:
    label_to_key = {_norm_header(label): key for key, label in VEHICLE_MASTER_COLUMNS}
    mapping: Dict[str, int] = {}
    for idx, h in enumerate(headers):
        key = label_to_key.get(_norm_header(h))
        if key:
            mapping[key] = idx
    return mapping


def _cell(row: tuple, idx: int) -> Any:
    if idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val).strip() if isinstance(val, str) else val


def parse_vehicle_master_bulk_upload(content: bytes) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ValueError("Worksheet is empty") from None

    col_map = _header_map(list(header_row))
    required = {"registration_number", "make", "model"}
    missing = required - set(col_map)
    if missing:
        raise ValueError(f"Template missing required columns: {', '.join(sorted(missing))}")

    parsed: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    for row_num, row in enumerate(rows_iter, start=DATA_START):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        record: Dict[str, Any] = {"row": row_num}
        for key, idx in col_map.items():
            raw = _cell(row, idx)
            if raw is None or raw == "":
                continue
            if key in _NUMERIC_KEYS:
                try:
                    record[key] = float(raw) if key != "year_of_manufacture" else int(float(raw))
                except (TypeError, ValueError):
                    errors.append({"row": row_num, "message": f"Invalid number for {key}: {raw}"})
                    continue
            else:
                record[key] = raw
        if not record.get("registration_number"):
            errors.append({"row": row_num, "message": "Registration Number is required"})
            continue
        parsed.append(record)
    return parsed, errors
