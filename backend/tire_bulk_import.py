"""Excel template and parsing for bulk tire import (upsert by serial number)."""
from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from country_utils import normalize_country_code
from models.enums import CurrencyEnum, TirePosition, TireStatus

BULK_HEADERS = [
    "Serial Number",
    "Brand",
    "Model",
    "Size",
    "Country",
    "Purchase Date",
    "Purchase Cost",
    "Currency",
    "Registration Number",
    "Position",
    "Mileage at Install",
    "Tread Depth (mm)",
    "Notes",
    "Status",
]

_HEADER_ALIASES = {
    "serial number": "serial_number",
    "serial": "serial_number",
    "brand": "brand",
    "model": "model",
    "size": "size",
    "country": "country",
    "purchase date": "purchase_date",
    "purchase cost": "purchase_cost",
    "cost": "purchase_cost",
    "currency": "currency",
    "registration number": "registration_number",
    "vehicle registration": "registration_number",
    "vehicle": "registration_number",
    "position": "position",
    "mileage at install": "mileage_at_install",
    "mileage": "mileage_at_install",
    "tread depth (mm)": "tread_depth_mm",
    "tread depth": "tread_depth_mm",
    "tread_depth_mm": "tread_depth_mm",
    "notes": "notes",
    "status": "status",
}

SAMPLE_ROW = [
    "TYRE-SN-001234",
    "Michelin",
    "XZY3",
    "265/70R17",
    "GH",
    "2026-01-15",
    450.0,
    "GHS",
    "GR-1234-20",
    "FRONT_LEFT",
    42000,
    8.5,
    "New install",
    "IN_USE",
]

_NUMBERED_INSTRUCTION = re.compile(r"^\d+\.\s")


def build_tire_template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tires"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")

    for col, title in enumerate(BULK_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font

    for col, value in enumerate(SAMPLE_ROW, start=1):
        ws.cell(row=2, column=col, value=value)

    for col in range(1, len(BULK_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    guide = wb.create_sheet("Instructions")
    guide["A1"] = "Tire bulk upload — how to use"
    guide["A1"].font = Font(bold=True)
    guide["A3"] = "1. Keep row 1 headers unchanged. Add one tire per row from row 3."
    guide["A4"] = "2. Serial Number is the unique key — existing tires are updated (upsert)."
    guide["A5"] = "3. Country: ISO alpha-2 (GH, LR, ST, …). Blank rows use the country selected in the upload dialog."
    guide["A6"] = "4. Registration Number matches an existing vehicle; leave blank for spare tires."
    guide["A7"] = "5. Position: FRONT_LEFT, FRONT_RIGHT, REAR_LEFT, REAR_RIGHT, or SPARE."
    guide["A8"] = "6. Status (optional, useful on update): IN_USE, SPARE, REPLACED, DISPOSED."
    guide["A9"] = "7. Dates: YYYY-MM-DD. Currency: GHS, LRD, USD, or STN."
    guide["A10"] = "8. Delete the sample row before uploading."
    guide.column_dimensions["A"].width = 88

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _normalize_header(value: Any) -> Optional[str]:
    if value is None:
        return None
    return _HEADER_ALIASES.get(str(value).strip().lower())


def _parse_date(value: Any, field: str = "Purchase Date") -> datetime:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        raise ValueError(f"{field} is required")
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if hasattr(value, "year") and hasattr(value, "month") and not hasattr(value, "hour"):
        return datetime(value.year, value.month, value.day)
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError as exc:
        raise ValueError(f"Invalid {field}: {value}") from exc


def _parse_float(value: Any, field: str, *, required: bool = True) -> Optional[float]:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        if required:
            raise ValueError(f"{field} is required")
        return None
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid {field}: {value}") from exc


def _normalize_currency(value: Any) -> str:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        raise ValueError("Currency is required")
    text = str(value).strip().upper()
    valid = {c.value for c in CurrencyEnum}
    if text not in valid:
        raise ValueError(f"Currency must be one of: {', '.join(sorted(valid))}")
    return text


def _normalize_position(value: Any) -> Optional[str]:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        return None
    text = str(value).strip().upper().replace(" ", "_").replace("-", "_")
    aliases = {
        "FL": TirePosition.FRONT_LEFT.value,
        "FR": TirePosition.FRONT_RIGHT.value,
        "RL": TirePosition.REAR_LEFT.value,
        "RR": TirePosition.REAR_RIGHT.value,
        "FRONTLEFT": TirePosition.FRONT_LEFT.value,
        "FRONTRIGHT": TirePosition.FRONT_RIGHT.value,
        "REARLEFT": TirePosition.REAR_LEFT.value,
        "REARRIGHT": TirePosition.REAR_RIGHT.value,
    }
    valid = {p.value for p in TirePosition}
    if text in valid:
        return text
    compact = text.replace("_", "")
    if compact in aliases:
        return aliases[compact]
    raise ValueError(f"Unknown Position: {value}")


def _normalize_status(value: Any) -> Optional[str]:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        return None
    text = str(value).strip().upper().replace(" ", "_").replace("-", "_")
    valid = {s.value for s in TireStatus}
    if text in valid:
        return text
    raise ValueError(f"Unknown Status: {value}")


def _row_is_empty(values: List[Any]) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in values)


def _cell_value(row: tuple, column_map: Dict[int, str], field: str) -> Any:
    for col_idx, name in column_map.items():
        if name == field:
            val = row[col_idx] if col_idx < len(row) else None
            if isinstance(val, str):
                return val.strip()
            return val
    return None


def _should_skip_row(row: tuple, column_map: Dict[int, str]) -> bool:
    if _row_is_empty(list(row)):
        return True
    serial = _cell_value(row, column_map, "serial_number")
    if serial is None or (isinstance(serial, str) and not str(serial).strip()):
        return True
    text = str(serial).strip()
    if text.lower().startswith("instructions") or _NUMBERED_INSTRUCTION.match(text):
        return True
    return False


def parse_tire_bulk_upload(
    file_bytes: bytes,
    *,
    default_country: Optional[str] = None,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb["Tires"] if "Tires" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The spreadsheet is empty")

    column_map: Dict[int, str] = {}
    for idx, header in enumerate(rows[0]):
        field = _normalize_header(header)
        if field:
            column_map[idx] = field

    present = set(column_map.values())
    required = {
        "serial_number",
        "brand",
        "model",
        "size",
        "purchase_date",
        "purchase_cost",
        "currency",
    }
    missing = [
        h for h in BULK_HEADERS
        if _HEADER_ALIASES.get(h.lower()) in required
        and _HEADER_ALIASES.get(h.lower()) not in present
    ]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    if "country" not in present and not default_country:
        raise ValueError(
            "Missing Country column and no default country provided. "
            "Add a Country column or select a country in the upload dialog."
        )

    parsed: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    seen_serials: set[str] = set()

    for row_idx, row in enumerate(rows[1:], start=2):
        if _should_skip_row(row, column_map):
            continue

        data: Dict[str, Any] = {"row": row_idx}
        try:
            for col_idx, field in column_map.items():
                val = row[col_idx] if col_idx < len(row) else None
                if isinstance(val, str):
                    val = val.strip()
                data[field] = val

            serial = str(data.get("serial_number") or "").strip()
            if not serial:
                raise ValueError("Serial Number is required")

            serial_key = serial.upper()
            if serial_key in seen_serials:
                raise ValueError(f"Duplicate serial number in file: {serial}")
            seen_serials.add(serial_key)

            brand = str(data.get("brand") or "").strip()
            model = str(data.get("model") or "").strip()
            size = str(data.get("size") or "").strip()
            if not brand:
                raise ValueError("Brand is required")
            if not model:
                raise ValueError("Model is required")
            if not size:
                raise ValueError("Size is required")

            country_raw = data.get("country")
            if country_raw is None or (isinstance(country_raw, str) and not str(country_raw).strip()):
                if not default_country:
                    raise ValueError("Country is required")
                country = normalize_country_code(default_country)
            else:
                country = normalize_country_code(str(country_raw).strip())

            reg_raw = data.get("registration_number")
            registration = str(reg_raw).strip() if reg_raw is not None and str(reg_raw).strip() else None

            notes_raw = data.get("notes")
            notes = str(notes_raw).strip() if notes_raw is not None and str(notes_raw).strip() else None

            parsed.append({
                "row": row_idx,
                "serial_number": serial,
                "brand": brand,
                "model": model,
                "size": size,
                "country": country,
                "purchase_date": _parse_date(data.get("purchase_date")),
                "purchase_cost": _parse_float(data.get("purchase_cost"), "Purchase Cost"),
                "currency": _normalize_currency(data.get("currency")),
                "registration_number": registration,
                "position": _normalize_position(data.get("position")),
                "mileage_at_install": _parse_float(
                    data.get("mileage_at_install"), "Mileage at Install", required=False
                ),
                "tread_depth_mm": _parse_float(
                    data.get("tread_depth_mm"), "Tread Depth (mm)", required=False
                ),
                "notes": notes,
                "status": _normalize_status(data.get("status")),
            })
        except ValueError as exc:
            errors.append({
                "row": row_idx,
                "serial_number": str(data.get("serial_number") or ""),
                "message": str(exc),
            })

    wb.close()
    return parsed, errors
